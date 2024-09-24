import requests
from bs4 import BeautifulSoup
import os
import openpyxl
import json
import re
def scrape_website(url):
    swedish_months = ["januari", "februari", "mars", "april", "maj", "juni", "juli", "augusti", "september", "oktober", "november", "december"]
    def add_zero_before_one_digit(array):
      modified_array = []
      for element in array:
        if element.startswith('P') or element.startswith('F'):
          if len(element[1:]) == 1:
            modified_array.append(f"{element[0]}0{element[1:]}")
          else:
            modified_array.append(element.strip())
        else:
          modified_array.append(element.strip())
      return modified_array
    def extract_month(date):
      
      extracted_month= ""
      for month in swedish_months:
        if month in date:
           
            extracted_month = month
            break
      return extracted_month.capitalize()
    def strip_date(date):
        new_date = date
        for index, month in enumerate(swedish_months):
          if month in new_date.lower():
           
            new_date = new_date.lower().replace(" ", "").replace(month, "/" + str(index +1))
        new_date = new_date.replace("2024", "").strip()
        return new_date

    try:
        """  # Fetch the website content
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
        response.raise_for_status()  # Raise exception for HTTP errors

        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'lxml')

        # Find the Excel link
        excel_link = soup.find('a', href=lambda href: href and href.lower().find('sanktionerade-turneringar-2024') != -1 and href.lower().endswith(('.xls', '.xlsx')))

        if not excel_link:
            print("No Excel file found.")
            return

        # Construct the full Excel URL
        excel_url = f"https://www.gbgfotboll.se//{excel_link['href']}"

        # Download the Excel file
        excel_response = requests.get(excel_url, headers={'User-Agent': 'Mozilla/5.0'})
        excel_response.raise_for_status()  # Raise exception for download errors

        # Create the download directory if it doesn't exist
        os.makedirs('./excel', exist_ok=True)  # Safe creation of directory

        # Generate a unique filename with underscores
        filename = os.path.basename(excel_url).replace('/', '_')
        # Change extension if needed
        filename = f"{filename.rsplit('.', 1)[0]}.xlsx"

        # Download the Excel content
        with open(f'./excel/{filename}', 'wb') as f:
            f.write(excel_response.content)
        # Load the Excel file using openpyxl
        workbook = openpyxl.load_workbook(f'./excel/{filename}')"""
        workbook = openpyxl.load_workbook(f'./excel/sanktionerade-turneringar-2024.xlsx')
        # Get the first worksheet (adjust index if needed)
        sheet = workbook.worksheets[0]
        # Add the "county" column to the header row

       

        header_row = sheet[1]
        new_cell = sheet.cell(row=1, column=len(header_row) + 1, value="county")
        header_row = list(header_row) + [new_cell]
        new_cell = sheet.cell(row=1, column=len(header_row) + 1, value="year")
        header_row = list(header_row) + [new_cell]
        new_cell = sheet.cell(row=1, column=len(header_row) + 1, value="categories")
        header_row = list(header_row) + [new_cell]
        new_cell = sheet.cell(row=1, column=len(header_row) + 1, value="month")
        header_row = list(header_row) + [new_cell]
        for cell in header_row:
                    if cell.value == "Hemsida":
                        cell.value = "link"
                    if cell.value == "Arrangör":
                        cell.value = "club"  
                    if cell.value == "Turneringar 2024":
                        cell.value = "name"  
                    if cell.value == "Åldersgrupp":
                        cell.value = "categoriesSummary"
                    if cell.value ==  "Spelform":
                        cell.value = "pitchSize"
                    if cell.value ==  "Datum":
                        cell.value = "date"    
                       
                       
                        
        
        # Add "Göteborg" to the "county" column for each row
        data = []
        for row in sheet.iter_rows(values_only=True, min_row=2):
            
            row_data = dict(zip([cell.value for cell in header_row], row))
            row_data["county"] = "Göteborg"
            row_data["year"] = "2024"
            row_data["categories"] = add_zero_before_one_digit(row_data["categoriesSummary"].split(','))
            row_data["month"] = extract_month(row_data["date"])

            row_data["date"] = strip_date(row_data["date"])
            data.append(row_data)
              
            # Save the data as JSON
        with open("sanktionerade-turneringar-2024.xlsx.json", 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        print(f"Excel file downloaded and converted to JSON successfully: sanktionerade-turneringar-2024.xlsx")

    except requests.exceptions.RequestException as e:
        print(f"Error fetching or downloading files: {e}")
    except Exception as e:
        print(f"Error processing Excel file: {e}")

if __name__ == "__main__":
    url = "https://www.gbgfotboll.se/tavling/fotboll/sanktionerade-turnering/"
    scrape_website(url)